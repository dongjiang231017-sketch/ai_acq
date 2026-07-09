import re
from io import BytesIO

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.api.auth import get_current_user
from app.db.session import get_db
from app.models.lead import MerchantLead
from app.models.user import User
from app.schemas.common import Page, paginate
from app.schemas.lead import LeadCreate, LeadRead

router = APIRouter()

# Excel 导入的表头映射（需求 7.5.2 MVP 必做：Excel 导入商家名单）
_IMPORT_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "name": ("商家名称", "名称", "店名", "门店名称", "商户名称", "name"),
    "phone": ("电话", "手机", "手机号", "联系电话", "公开电话", "电话号码", "phone"),
    "city": ("城市", "所在城市", "city"),
    "category": ("类目", "分类", "行业", "category"),
    "address": ("地址", "门店地址", "详细地址", "address"),
    "district": ("区县", "区域", "district"),
    "platform": ("平台", "来源", "平台来源", "platform", "source"),
}


def _normalize_import_phone(value: object) -> str | None:
    text = str(value or "").strip()
    # Excel 数字单元格会读成浮点（"13800000000.0"）：先剥掉小数部分，
    # 否则去符号后变成 12 位错号还能通过座机校验入库
    if re.fullmatch(r"\d+\.0+", text):
        text = text.split(".", 1)[0]
    digits = re.sub(r"\D+", "", text)
    if digits.startswith("86") and len(digits) == 13:
        digits = digits[2:]
    if len(digits) == 11 and digits.startswith("1"):
        return digits
    if 7 <= len(digits) <= 12 and not digits.startswith("1"):  # 座机（含区号）
        return digits
    return None


@router.post("/import-excel")
async def import_leads_excel(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict[str, object]:
    """Excel 导入商家名单：识别表头、归一化电话、按电话/店名+地址去重后入库。"""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise HTTPException(status_code=500, detail="服务器缺少 openpyxl 依赖") from exc

    raw = await file.read()
    if len(raw) > 20 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="文件超过 20MB 限制")
    try:
        workbook = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"无法解析 Excel 文件：{exc}") from exc

    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        header = [str(cell or "").strip() for cell in next(rows)]
    except StopIteration:
        raise HTTPException(status_code=400, detail="Excel 内容为空") from None

    column_map: dict[str, int] = {}
    for field, aliases in _IMPORT_HEADER_ALIASES.items():
        for index, title in enumerate(header):
            if title and any(alias.lower() == title.lower() for alias in aliases):
                column_map[field] = index
                break
    if "name" not in column_map:
        raise HTTPException(status_code=400, detail=f"未找到「商家名称」列，实际表头：{header[:10]}")

    # 去重基础：现有线索的归一化电话 + 店名+地址（一次查询预热，避免逐条查库）
    existing_phones: set[str] = set()
    existing_name_addr: set[tuple[str, str]] = set()
    for lead_phone, lead_name, lead_address in db.execute(
        select(MerchantLead.phone, MerchantLead.name, MerchantLead.address).where(
            MerchantLead.owner_user_id == current_user.id,
        ),
    ).all():
        normalized = _normalize_import_phone(lead_phone)
        if normalized:
            existing_phones.add(normalized)
        if lead_name and lead_address:
            existing_name_addr.add((str(lead_name), str(lead_address)))

    def _cell(row: tuple, field: str) -> str:
        index = column_map.get(field)
        if index is None or index >= len(row):
            return ""
        return str(row[index] or "").strip()

    total = inserted = duplicated = invalid = 0
    for row in rows:
        if row is None or all(cell in (None, "") for cell in row):
            continue
        total += 1
        name = _cell(row, "name")
        phone = _normalize_import_phone(_cell(row, "phone"))
        address = _cell(row, "address") or None
        if not name:
            invalid += 1
            continue
        if not phone:
            invalid += 1
            continue
        if phone in existing_phones or (address and (name, address) in existing_name_addr):
            duplicated += 1
            continue
        db.add(
            MerchantLead(
                name=name,
                platform=_cell(row, "platform") or "导入",
                city=_cell(row, "city") or "未知",
                category=_cell(row, "category") or "未分类",
                phone=phone,
                district=_cell(row, "district") or None,
                address=address,
                source="Excel导入",
                intent_score=60,
                status="待外呼",
                follow_up_status="未跟进",
                remark=f"Excel导入：{file.filename or ''}",
                owner_user_id=current_user.id,
                created_by_user_id=current_user.id,
            ),
        )
        existing_phones.add(phone)
        if address:
            existing_name_addr.add((name, address))
        inserted += 1

    db.commit()
    return {"total": total, "inserted": inserted, "duplicated": duplicated, "invalid": invalid}


@router.get("", response_model=Page[LeadRead])
def list_leads(
    source: str | None = Query(default=None),
    platform: str | None = Query(default=None),
    city: str | None = Query(default=None),
    category: str | None = Query(default=None),
    status: str | None = Query(default=None),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=1000, ge=1, le=1000, alias="pageSize"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> Page[LeadRead]:
    statement = select(MerchantLead).where(
        MerchantLead.owner_user_id == current_user.id,
        MerchantLead.phone.is_not(None),
        MerchantLead.phone.not_in(["", "-"]),
    )
    if source:
        statement = statement.where(MerchantLead.source == source)
    if platform:
        statement = statement.where(MerchantLead.platform == platform)
    if city:
        statement = statement.where(MerchantLead.city.ilike(f"%{city}%"))
    if category:
        statement = statement.where(MerchantLead.category.ilike(f"%{category}%"))
    if status:
        statement = statement.where(MerchantLead.status == status)
    total = db.scalar(select(func.count()).select_from(statement.subquery())) or 0
    offset, page = paginate(total, page, page_size)
    items = db.scalars(statement.order_by(MerchantLead.created_at.desc()).offset(offset).limit(page_size)).all()
    return Page(
        items=[LeadRead.model_validate(i, from_attributes=True) for i in items],
        total=total,
        page=page,
        pageSize=page_size,
    )


@router.post("", response_model=LeadRead)
def create_lead(
    payload: LeadCreate,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> MerchantLead:
    lead = MerchantLead(
        name=payload.name,
        platform=payload.platform,
        city=payload.city,
        category=payload.category,
        phone=payload.phone,
        contact_name=payload.contact_name,
        contact_title=payload.contact_title,
        wechat_id=payload.wechat_id,
        platform_url=payload.platform_url,
        platform_homepage_url=payload.platform_homepage_url or payload.platform_url,
        source_poi_id=payload.source_poi_id,
        province=payload.province,
        district=payload.district,
        address=payload.address,
        longitude=payload.longitude,
        latitude=payload.latitude,
        source=payload.source,
        intent_score=payload.intent_score,
        status=payload.status,
        follow_up_status=payload.follow_up_status,
        remark=payload.remark,
        owner_user_id=current_user.id,
        created_by_user_id=current_user.id,
    )
    db.add(lead)
    db.commit()
    db.refresh(lead)
    return lead
